"""Stage 2 — Consolidate all supplier files for a single segment.

Reads ``<output_base>/tmp/raw/<segment>/*.xlsx``, groups by FY sheet,
stacks all suppliers vertically (first supplier contributes the header row,
subsequent ones contribute only data rows), and writes
``<output_base>/tmp/segment/Consolidated_<segment>_<YYYYMMDD>.xlsx``.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

FY_SHEET_RE = re.compile(r"^FY\d{2}$", re.IGNORECASE)


def _is_all_none(row: tuple | list) -> bool:
    return all(v is None for v in row)


def consolidate_segment(
    segment_name: str,
    raw_dir: Path,
    output_base: Path,
    log: Callable[[str, str], None],
) -> Path | None:
    """Consolidate all supplier xlsx files for *segment_name* into one workbook.

    *raw_dir* is the folder containing the per-supplier .xlsx files.
    Returns the path of the saved file, or ``None`` on failure.
    """
    if not raw_dir.exists():
        log(f"  [{segment_name}] raw dir not found: {raw_dir}", "ERROR")
        return None

    xlsx_files = sorted(f for f in raw_dir.glob("*.xlsx") if not f.name.startswith("~$"))
    if not xlsx_files:
        log(f"  [{segment_name}] no .xlsx files in {raw_dir}", "WARNING")
        return None

    # -----------------------------------------------------------------------
    # Discover which FY sheets exist across all supplier files
    # -----------------------------------------------------------------------
    all_fy: set[str] = set()
    for xlsx_path in xlsx_files:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            t = ws.title.strip().upper()
            if FY_SHEET_RE.match(t):
                all_fy.add(t)
        wb.close()

    if not all_fy:
        log(f"  [{segment_name}] no FY sheets found", "WARNING")
        return None

    out_wb: Workbook = Workbook()
    out_wb.remove(out_wb.active)  # type: ignore[arg-type]

    # -----------------------------------------------------------------------
    # Per-FY consolidation
    # -----------------------------------------------------------------------
    for fy in sorted(all_fy):
        out_ws = out_wb.create_sheet(fy)
        first_file = True
        valid_cols: list[int] = []
        row_count = 0

        for xlsx_path in xlsx_files:
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
            ws = None
            for sheet in wb.worksheets:
                if sheet.title.strip().upper() == fy:
                    ws = sheet
                    break

            if ws is None:
                wb.close()
                continue

            # rows[0] = title row (skip), rows[1] = column headers, rows[2:] = data
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                continue  # no header row available

            header_row = rows[1]  # source Row 2

            if first_file:
                # Determine valid columns (non-empty header)
                valid_cols = [
                    i for i, v in enumerate(header_row)
                    if v is not None and str(v).strip()
                ]
                if not valid_cols:
                    continue
                # Write normalised header (replace \n with space)
                out_ws.append([
                    str(header_row[i]).replace("\n", " ").replace("\r", " ").strip()
                    if header_row[i] is not None else None
                    for i in valid_cols
                ])
                first_file = False

            # Write data rows (source Row 3+)
            for row in rows[2:]:
                if _is_all_none(row):
                    continue
                filtered = [row[i] if i < len(row) else None for i in valid_cols]
                if not _is_all_none(filtered):
                    out_ws.append(filtered)
                    row_count += 1

        log(f"  [{segment_name}] {fy}: {row_count} data rows from {len(xlsx_files)} supplier(s)", "INFO")

    seg_dir = output_base / "tmp" / "segment"
    seg_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    out_path = seg_dir / f"Consolidated_{segment_name}_{date_str}.xlsx"
    out_wb.save(out_path)
    log(f"  [{segment_name}] saved → {out_path.name}", "INFO")
    return out_path

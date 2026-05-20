"""Stage 3 — Merge all segment Consolidated files into one workbook.

Reads the four ``Consolidated_<segment>_<YYYYMMDD>.xlsx`` files produced by
Stage 2 and stacks them vertically per FY sheet.  The first segment in each
sheet contributes the header row; subsequent segments contribute only data.

Output: ``<output_base>/tmp/all/consolidate_all_<YYYYMMDD>.xlsx``
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _is_all_none(row: tuple | list) -> bool:
    return all(v is None for v in row)


def consolidate_all(
    segment_paths: list[Path],
    output_base: Path,
    log: Callable[[str, str], None],
) -> Path | None:
    """Merge all segment consolidated files into a single workbook.

    Returns the path of the saved file, or ``None`` on failure.
    """
    if not segment_paths:
        log("  [All] no segment files to merge", "ERROR")
        return None

    # -----------------------------------------------------------------------
    # Discover all FY sheets across all segment files
    # -----------------------------------------------------------------------
    all_fy: set[str] = set()
    for p in segment_paths:
        wb = load_workbook(p, data_only=True, read_only=True)
        for ws in wb.worksheets:
            all_fy.add(ws.title.strip().upper())
        wb.close()

    out_wb: Workbook = Workbook()
    out_wb.remove(out_wb.active)  # type: ignore[arg-type]

    # -----------------------------------------------------------------------
    # Per-FY merge
    # -----------------------------------------------------------------------
    for fy in sorted(all_fy):
        out_ws = out_wb.create_sheet(fy)
        first_seg = True
        merged_segs = 0

        for seg_path in segment_paths:
            wb = load_workbook(seg_path, data_only=True, read_only=True)
            ws = None
            for sheet in wb.worksheets:
                if sheet.title.strip().upper() == fy:
                    ws = sheet
                    break

            if ws is None:
                wb.close()
                continue

            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                continue

            if first_seg:
                for row in rows:
                    if not _is_all_none(row):
                        out_ws.append(list(row))
                first_seg = False
            else:
                # Skip first row (header already written)
                for row in rows[1:]:
                    if not _is_all_none(row):
                        out_ws.append(list(row))

            merged_segs += 1

        log(f"  [All] {fy}: merged {merged_segs} segment(s)", "INFO")

    all_dir = output_base / "tmp" / "all"
    all_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    out_path = all_dir / f"consolidate_all_{date_str}.xlsx"
    out_wb.save(out_path)
    log(f"  [All] saved → {out_path.name}", "INFO")
    return out_path

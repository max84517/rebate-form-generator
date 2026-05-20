"""Stage 4 — Remove confidential cost columns (HP Cost, ODM Cost).

Reads the consolidated-all workbook and strips any column whose header
contains "HP Cost" or "ODM Cost" (case-insensitive).

Output: ``<output_base>/tmp/rebate/rebate_only.xlsx``
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

_COST_KEYWORDS = ("hp cost", "odm cost")


def _is_cost_header(value: object) -> bool:
    if value is None:
        return False
    normalised = str(value).replace("\n", " ").replace("\r", " ").lower()
    return any(kw in normalised for kw in _COST_KEYWORDS)


def build_rebate_only(
    all_path: Path | None,
    output_base: Path,
    log: Callable[[str, str], None],
) -> Path | None:
    """Strip HP Cost / ODM Cost columns from *all_path*.

    Returns the path of the saved file, or ``None`` on failure.
    """
    if all_path is None or not all_path.exists():
        log(f"  [Rebate] input file not found: {all_path}", "ERROR")
        return None

    wb_in = load_workbook(all_path, data_only=True)
    wb_out: Workbook = Workbook()
    wb_out.remove(wb_out.active)  # type: ignore[arg-type]

    for ws in wb_in.worksheets:
        ws_out = wb_out.create_sheet(ws.title)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row = rows[0]
        keep_cols = [i for i, v in enumerate(header_row) if not _is_cost_header(v)]
        removed = len(header_row) - len(keep_cols)

        for row in rows:
            filtered = [row[i] if i < len(row) else None for i in keep_cols]
            ws_out.append(filtered)

        log(
            f"  [Rebate] {ws.title}: kept {len(keep_cols)} cols, removed {removed} cost col(s)",
            "INFO",
        )

    rebate_dir = output_base / "tmp" / "rebate"
    rebate_dir.mkdir(parents=True, exist_ok=True)
    out_path = rebate_dir / "rebate_only.xlsx"
    wb_out.save(out_path)
    log(f"  [Rebate] saved → {out_path.name}", "INFO")
    return out_path

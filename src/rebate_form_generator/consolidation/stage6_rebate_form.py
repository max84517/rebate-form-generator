"""Stage 6 — Generate rebate form input.xlsx from rebate raw.xlsx.

Quarter rules
-------------
Q1 : Nov (prev year), Dec (prev year), Jan  (FY year)
Q2 : Feb, Mar, Apr  (FY year)
Q3 : May, Jun, Jul  (FY year)
Q4 : Aug, Sep, Oct  (FY year)

Example: FY26 Q1 → Nov 2025, Dec 2025, Jan 2026
"""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook, Workbook

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

REBATE_COL_RE = re.compile(
    r"^Rebate\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})$",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Quarter helpers
# ---------------------------------------------------------------------------

def quarter_months(fy: int, q: int) -> list[tuple[int, int]]:
    """Return [(month, year), ...] for the 3 months of FY<fy> Q<q>.

    *fy* is the 2-digit year (e.g. 26 for FY26).
    """
    full_year = 2000 + fy
    if q == 1:
        return [(11, full_year - 1), (12, full_year - 1), (1, full_year)]
    elif q == 2:
        return [(2, full_year), (3, full_year), (4, full_year)]
    elif q == 3:
        return [(5, full_year), (6, full_year), (7, full_year)]
    else:  # q == 4
        return [(8, full_year), (9, full_year), (10, full_year)]


def current_fy_quarter() -> tuple[int, int]:
    """Return (fy_2digit, quarter) based on today's date."""
    today = date.today()
    m, y = today.month, today.year
    if m in (11, 12):
        return (y + 1) % 100, 1
    elif m == 1:
        return y % 100, 1
    elif m in (2, 3, 4):
        return y % 100, 2
    elif m in (5, 6, 7):
        return y % 100, 3
    else:  # m in (8, 9, 10)
        return y % 100, 4


def _col_header(month: int, year: int) -> str:
    return f"Rebate {MONTH_ABBR[month - 1]} {year}"


# ---------------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------------

def generate_rebate_form(
    rebate_raw_path: Path,
    fy: int,
    quarter: int,
    selected_columns: list[str],
    output_dir: Path,
    log: Callable[[str, str], None],
) -> list[Path]:
    """Read rebate raw.xlsx and write per-supplier contract input files.

    For each data row the function emits 1–3 output rows depending on whether
    the rebate price changes across the three months of the quarter.
    Rows are split by GTK Suppliers and saved as
    ``contract input - <Supplier>.xlsx``.

    Returns a list of paths to the written files (empty list on failure).
    """
    if not rebate_raw_path.exists():
        log(f"rebate raw.xlsx not found: {rebate_raw_path}", "ERROR")
        return []

    wb_in = load_workbook(rebate_raw_path, data_only=True)
    ws_in = wb_in.active

    months = quarter_months(fy, quarter)          # [(month, year), ...]
    target_headers = [_col_header(m, y) for m, y in months]
    start_dates = [date(y, m, 1) for m, y in months]

    log(f"  Quarter months: {target_headers}", "INFO")

    # ── Locate columns in header row ─────────────────────────────────────
    all_headers: list = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    header_lower_map: dict[str, int] = {}
    for i, h in enumerate(all_headers):
        if h is not None:
            header_lower_map[str(h).strip().lower()] = i

    target_col_indices: list[int | None] = []
    for h in target_headers:
        idx = header_lower_map.get(h.lower())
        if idx is None:
            log(f"  Column '{h}' not found in rebate raw.xlsx — will treat as None", "WARNING")
        target_col_indices.append(idx)

    # ── Build feature columns: selected optional + GTK Suppliers ─────────
    gtk_col_idx: int | None = header_lower_map.get("gtk suppliers")

    feature_col_indices: list[int] = []
    feature_headers: list[str] = []
    for col_name in selected_columns:
        idx = header_lower_map.get(col_name.lower())
        if idx is not None:
            feature_col_indices.append(idx)
            feature_headers.append(all_headers[idx])
        else:
            log(f"  Optional column '{col_name}' not found — skipping", "WARNING")

    if gtk_col_idx is not None:
        feature_col_indices.append(gtk_col_idx)
        feature_headers.append(all_headers[gtk_col_idx])
    else:
        log("  'GTK Suppliers' column not found in data", "WARNING")

    # Index of the supplier value inside each output row tuple
    supplier_tuple_idx: int | None = (
        len(feature_col_indices) - 1 if gtk_col_idx is not None else None
    )

    out_headers = feature_headers + ["Per-Unit Rebate Amount $USD", "Rebate Period Start Date"]
    price_col = len(feature_headers) + 1  # 1-based

    # ── Collect all rows ──────────────────────────────────────────────────
    all_rows: list[tuple] = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        feature_vals = [row[i] if i < len(row) else None for i in feature_col_indices]

        prices = [
            (row[i] if i < len(row) else None) if i is not None else None
            for i in target_col_indices
        ]

        # Build price segments: start a new segment whenever the price changes
        segments: list[tuple] = []
        for price, seg_date in zip(prices, start_dates):
            if segments and segments[-1][0] == price:
                continue  # same price — extend current segment
            segments.append((price, seg_date))

        for price, seg_date in segments:
            rounded_price = round(float(price), 2) if price is not None else 0.0
            all_rows.append(tuple(feature_vals + [rounded_price, seg_date]))

    # ── Deduplicate while preserving order ───────────────────────────────
    seen: set[tuple] = set()
    unique_rows: list[tuple] = []
    for row_tuple in all_rows:
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique_rows.append(row_tuple)

    dropped = len(all_rows) - len(unique_rows)
    if dropped:
        log(f"  Dropped {dropped} duplicate row(s)", "INFO")

    # ── Group by supplier ─────────────────────────────────────────────────
    supplier_rows: dict[str, list[tuple]] = defaultdict(list)
    for row_tuple in unique_rows:
        if supplier_tuple_idx is not None:
            supplier = str(row_tuple[supplier_tuple_idx] or "Unknown").strip()
        else:
            supplier = "Unknown"
        supplier_rows[supplier].append(row_tuple)

    # ── Write one file per supplier ───────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    out_paths: list[Path] = []
    for supplier, rows in sorted(supplier_rows.items()):
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Input"
        ws_out.append(out_headers)
        for i, row_tuple in enumerate(rows, start=2):
            ws_out.append(list(row_tuple))
            ws_out.cell(row=i, column=price_col).number_format = '"$"#,##0.00'
        out_path = output_dir / f"contract input - {supplier}.xlsx"
        wb_out.save(out_path)
        log(f"  Saved {len(rows)} rows → {out_path.name}", "INFO")
        out_paths.append(out_path)

    return out_paths

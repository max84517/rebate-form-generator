"""Stage 5 — Write the final Pricing Template Excel.

Reads the rebate-only workbook, takes the user-selected FY sheet,
filters out rows where the ``Platforms/Project`` column is empty,
and writes a single-sheet workbook to
``<output_base>/pricing_template/Pricing_Template_InputDevices.xlsx``.
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _normalise_header(value: object) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\n", " ")
        .replace("\r", " ")
        .strip()
        .lower()
    )


def write_pricing_template(
    rebate_path: Path | None,
    fy_sheet: str,
    output_base: Path,
    log: Callable[[str, str], None],
) -> Path | None:
    """Write the pricing template for *fy_sheet* from *rebate_path*.

    Returns the output path, or ``None`` on failure.
    """
    if rebate_path is None or not rebate_path.exists():
        log(f"  [Template] input file not found: {rebate_path}", "ERROR")
        return None

    wb = load_workbook(rebate_path, data_only=True)

    # Find the requested FY sheet (case-insensitive)
    ws = None
    target = fy_sheet.strip().upper()
    for sheet in wb.worksheets:
        if sheet.title.strip().upper() == target:
            ws = sheet
            break

    if ws is None:
        log(f"  [Template] sheet '{fy_sheet}' not found", "ERROR")
        available = [s.title for s in wb.worksheets]
        log(f"  [Template] available sheets: {available}", "INFO")
        return None

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        log("  [Template] sheet is empty", "ERROR")
        return None

    header_row = rows[0]

    # Locate Platforms/Project column
    platforms_col: int | None = None
    for i, v in enumerate(header_row):
        if _normalise_header(v) == "platforms/project":
            platforms_col = i
            break

    if platforms_col is None:
        log("  [Template] 'Platforms/Project' column not found — writing all rows", "WARNING")

    # Build output workbook
    out_wb: Workbook = Workbook()
    out_ws = out_wb.active
    out_ws.title = target  # type: ignore[union-attr]

    out_ws.append(list(header_row))  # type: ignore[union-attr]

    data_count = 0
    for row in rows[1:]:
        if platforms_col is not None:
            val = row[platforms_col] if platforms_col < len(row) else None
            if val is None or str(val).strip() == "":
                continue
        out_ws.append(list(row))  # type: ignore[union-attr]
        data_count += 1

    out_dir = output_base
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Pricing_Template_InputDevices.xlsx"
    out_wb.save(out_path)
    log(f"  [Template] saved → {out_path}  ({data_count} data rows)", "INFO")
    return out_path

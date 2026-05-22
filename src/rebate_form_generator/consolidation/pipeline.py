"""Public pipeline API.

Usage::

    from rebate_form_generator.consolidation.pipeline import (
        get_available_fy_sheets,
        run_full_pipeline,
    )

    source_paths = {
        "nb_kb":      r"C:\\Data\\NB KB",
        "dt_kb":      r"C:\\Data\\DT KB",
        "peripheral": r"C:\\Data\\Peripheral",
    }
    output_path = Path("data/output")

    def log(msg, level):
        print(f"[{level}] {msg}")

    # Run Stage 1-4, get available FY sheets
    fy_list = get_available_fy_sheets(source_paths, output_path, log)

    # Run Stage 5 (reuses cached Stage 1-4 results)
    out = run_full_pipeline(source_paths, "FY25", output_path, log)
"""
from __future__ import annotations

import os
import shutil
import stat
import tempfile
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from .stage1_ingest import ingest_nb, ingest_segment
from .stage2_segment import consolidate_segment
from .stage3_all import consolidate_all
from .stage4_rebate import build_rebate_only
from .stage5_template import write_pricing_template
from .stage6_rebate_form import generate_rebate_form
from .stage7_report import generate_report

# Module-level cache: stores results from the last successful Stage 1-4 run
_cache: dict = {
    "rebate_path": None,
    "output_path": None,
    "pricing_data_dir": None,
    "coverage": None,
}


def get_available_fy_sheets(
    source_paths: dict,
    output_path: Path,
    log: Callable[[str, str], None],
) -> list[str]:
    """Run Stages 1–4 and return the list of available FY sheet names.

    Results are cached so that a subsequent :func:`run_full_pipeline` call
    with the same *output_path* can skip Stages 1–4.
    """
    # Directories derived from output_path
    #   data/source data  → ingested supplier workbooks
    #   data/pricing data → final output (rebate raw.xlsx)
    #   system temp       → stage 2-4 intermediates (never written to data/)
    source_data_dir = output_path.parent / "source data"
    pricing_data_dir = output_path.parent / "rebate raw"

    def _rmtree(path: Path) -> None:
        def _on_err(func, p, exc_info):
            try:
                os.chmod(p, stat.S_IWRITE)
                func(p)
            except Exception:
                pass
        shutil.rmtree(path, onexc=_on_err)

    # Clear source data sub-dirs (NB, DT, Peripheral) before fresh ingest
    for seg_dir in (
        source_data_dir / "NB",
        source_data_dir / "DT",
        source_data_dir / "Peripheral",
    ):
        if seg_dir.exists():
            log(f"  Clearing {seg_dir.name} source data …", "INFO")
            _rmtree(seg_dir)

    # Remove legacy output/tmp if it still exists from an older run
    legacy_tmp = output_path / "tmp"
    if legacy_tmp.exists():
        _rmtree(legacy_tmp)

    # Stage 2-4 intermediates go to the system temp directory
    tmp_dir = Path(tempfile.gettempdir()) / "rebate_form_generator"
    if tmp_dir.exists():
        _rmtree(tmp_dir)
    tmp_dir.mkdir(parents=True, exist_ok=True)

    nb_kb = Path(source_paths["nb_kb"])
    dt_kb = Path(source_paths["dt_kb"])
    peripheral = Path(source_paths["peripheral"])

    # ------------------------------------------------------------------
    # Stage 1 — Ingest
    # ------------------------------------------------------------------
    log("=== Stage 1: Ingest ===", "INFO")
    log("Processing NB KB …", "INFO")
    nb_cov = ingest_nb(nb_kb, source_data_dir / "NB", log)

    log("Processing DT KB …", "INFO")
    dt_cov = ingest_segment(dt_kb, "DT", source_data_dir / "DT", log)

    log("Processing Peripheral …", "INFO")
    peri_cov = ingest_segment(peripheral, "Peripheral", source_data_dir / "Peripheral", log)

    # Build combined coverage: segment → {supplier: [FY sheets]}
    coverage: dict[str, dict[str, list[str]]] = {
        "bNB":        nb_cov.get("bNB", {}),
        "cNB":        nb_cov.get("cNB", {}),
        "DT":         dt_cov,
        "Peripheral": peri_cov,
    }

    # ------------------------------------------------------------------
    # Stage 2 — Consolidate by segment
    # ------------------------------------------------------------------
    log("=== Stage 2: Consolidate by Segment ===", "INFO")
    nb_raw_dir = source_data_dir / "NB"
    segment_files: list[Path] = []
    for seg, raw_dir, suffix in [
        ("bNB",        nb_raw_dir,                     "bNB"),
        ("cNB",        nb_raw_dir,                     "cNB"),
        ("DT",         source_data_dir / "DT",         None),
        ("Peripheral", source_data_dir / "Peripheral", None),
    ]:
        log(f"Consolidating {seg} …", "INFO")
        path = consolidate_segment(seg, raw_dir, tmp_dir, log, sheet_suffix=suffix)
        if path is not None:
            segment_files.append(path)

    # ------------------------------------------------------------------
    # Stage 3 — Consolidate all segments
    # ------------------------------------------------------------------
    log("=== Stage 3: Consolidate All Segments ===", "INFO")
    all_path = consolidate_all(segment_files, tmp_dir, log)

    # ------------------------------------------------------------------
    # Stage 4 — Rebate-only (remove cost columns)
    # ------------------------------------------------------------------
    log("=== Stage 4: Rebate Only ===", "INFO")
    rebate_path = build_rebate_only(all_path, tmp_dir, log)

    # Cache results
    _cache["rebate_path"] = rebate_path
    _cache["output_path"] = output_path
    _cache["pricing_data_dir"] = pricing_data_dir
    _cache["coverage"] = coverage

    if rebate_path is None:
        return [], {}

    wb = load_workbook(rebate_path, data_only=True, read_only=True)
    fy_sheets = [ws.title for ws in wb.worksheets]
    wb.close()

    log(f"Available FY sheets: {fy_sheets}", "INFO")
    return fy_sheets, coverage


def run_full_pipeline(
    source_paths: dict,
    fy_sheet: str,
    output_path: Path,
    log: Callable[[str, str], None],
) -> Path | None:
    """Run Stage 5 (and Stages 1–4 if the cache is stale).

    Returns the path to ``Pricing_Template_InputDevices.xlsx``, or ``None``.
    """
    rebate_path: Path | None = _cache.get("rebate_path")
    cached_output: Path | None = _cache.get("output_path")

    if rebate_path is None or cached_output != output_path:
        log("Cache miss — running Stages 1–4 first …", "INFO")
        get_available_fy_sheets(source_paths, output_path, log)
        rebate_path = _cache.get("rebate_path")

    pricing_data_dir: Path = _cache.get("pricing_data_dir") or (output_path.parent / "pricing data")

    log("=== Stage 5: Write Pricing Template ===", "INFO")
    return write_pricing_template(rebate_path, fy_sheet, pricing_data_dir, log)


def check_latest_files(
    source_paths: dict,
    log: Callable[[str, str], None],
) -> dict[str, dict[str, str]]:
    """Preview the latest .xlsx file chosen per supplier (no file copying)."""
    from .stage1_ingest import _find_segment_root, _get_latest_xlsx, _parse_supplier_name
    import re

    NB_ROOT_KW = "Master price table_NB"
    DT_ROOT_KW = "Master price table_DT"
    PERI_ROOT_KW = "Master price table_Peripheral"

    result: dict[str, dict[str, str]] = {}
    spec = [
        ("NB KB", Path(source_paths["nb_kb"]), NB_ROOT_KW),
        ("DT KB", Path(source_paths["dt_kb"]), DT_ROOT_KW),
        ("Peripheral", Path(source_paths["peripheral"]), PERI_ROOT_KW),
    ]
    for label, root, kw in spec:
        seg_root = _find_segment_root(root, kw)
        if seg_root is None:
            log(f"[{label}] segment root not found", "WARNING")
            continue
        result[label] = {}
        for supplier_folder in sorted(seg_root.iterdir()):
            if not supplier_folder.is_dir():
                continue
            latest = _get_latest_xlsx(supplier_folder)
            supplier = _parse_supplier_name(supplier_folder.name)
            result[label][supplier] = str(latest) if latest else "(no xlsx)"
            log(f"  [{label}] {supplier}: {latest.name if latest else 'missing'}", "INFO")
    return result


def check_missing_fy_sheets(
    source_paths: dict,
    fy: str,
    log: Callable[[str, str], None],
) -> list[str]:
    """Return supplier names that are missing the specified *fy* sheet."""
    from .stage1_ingest import _find_segment_root, _get_latest_xlsx, _parse_supplier_name
    import re

    FY_RE = re.compile(r"^FY\d{2}$", re.IGNORECASE)
    NB_SHEET_RE = re.compile(r"^FY(\d{2})\s*([cb])\s*NB$", re.IGNORECASE)
    missing: list[str] = []
    target = fy.strip().upper()

    spec = [
        (Path(source_paths["nb_kb"]), "Master price table_NB", True),
        (Path(source_paths["dt_kb"]), "Master price table_DT", False),
        (Path(source_paths["peripheral"]), "Master price table_Peripheral", False),
    ]
    for root, kw, is_nb in spec:
        seg_root = _find_segment_root(root, kw)
        if seg_root is None:
            continue
        for supplier_folder in sorted(seg_root.iterdir()):
            if not supplier_folder.is_dir():
                continue
            latest = _get_latest_xlsx(supplier_folder)
            if latest is None:
                continue
            supplier = _parse_supplier_name(supplier_folder.name)
            wb = load_workbook(latest, data_only=True, read_only=True)
            if is_nb:
                found = any(
                    f"FY{m.group(1)}" == target
                    for s in wb.worksheets
                    if (m := NB_SHEET_RE.match(s.title.strip()))
                )
            else:
                found = any(
                    s.title.strip().upper() == target for s in wb.worksheets
                )
            wb.close()
            if not found:
                missing.append(supplier)
                log(f"  MISSING {target}: {supplier}", "WARNING")
    return missing


def run_rebate_form_pipeline(
    output_path: Path,
    fy: int,
    quarter: int,
    selected_columns: list[str],
    log: Callable[[str, str], None],
) -> list[Path]:
    """Stage 6: generate per-supplier contract input files from ``rebate raw.xlsx``.

    *fy* is the 2-digit FY year (e.g. 26 for FY26).
    *quarter* is 1–4.
    *selected_columns* are the optional feature columns to keep.
    Returns a list of paths to the written files (empty list on failure).
    """
    rebate_raw_path = output_path.parent / "rebate raw" / "rebate raw.xlsx"
    output_dir = output_path.parent / "rebate form input"
    log(f"=== Stage 6: Generate Form Data (FY{fy:02d} Q{quarter}) ===", "INFO")
    return generate_rebate_form(rebate_raw_path, fy, quarter, selected_columns, output_dir, log)


def run_report_pipeline(
    output_path: Path,
    suppliers: list[str],
    form_number: str,
    log: Callable[[str, str], None],
) -> list[Path]:
    """Stage 7: generate per-supplier Word contracts.

    *suppliers* is the list of supplier names to process.
    *form_number* is the user-supplied Form# string.
    Returns a list of paths to the written .docx files.
    """
    base = output_path.parent
    supplier_info_dir = base / "supplier info"
    template_dir = base / "template"
    rebate_form_input_dir = base / "rebate form input"
    log(f"=== Stage 7: Generate Report (Form#{form_number}) ===", "INFO")
    return generate_report(
        supplier_info_dir, template_dir, rebate_form_input_dir,
        output_path, suppliers, form_number, log,
    )

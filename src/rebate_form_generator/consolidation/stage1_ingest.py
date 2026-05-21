"""Stage 1 — Ingest raw supplier Excel files.

Reads the latest .xlsx from each supplier sub-folder, applies the
GTK Suppliers fix, and writes split/normalised workbooks to
the caller-supplied *raw_dir*.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# ---------------------------------------------------------------------------
# Regexes
# ---------------------------------------------------------------------------
NB_SHEET_RE = re.compile(r"^FY(\d{2})\s*([cb])\s*NB$", re.IGNORECASE)
FY_SHEET_RE = re.compile(r"^FY\d{2}$", re.IGNORECASE)


def _make_supplier_folder_re(segment: str) -> re.Pattern:
    """Return a regex matching 'Master price table_<segment>_<Supplier>'."""
    return re.compile(
        rf"^Master\s+price\s+table_{re.escape(segment)}_.+$", re.IGNORECASE
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _get_latest_xlsx(folder: Path) -> Path | None:
    """Return the most recently modified .xlsx that is not an Office temp file."""
    files = [f for f in folder.glob("*.xlsx") if not f.name.startswith("~$")]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def _parse_supplier_name(folder_name: str) -> str:
    """Extract supplier from folder name.

    ``"Master price table_NB_Chicony"``  →  ``"Chicony"``
    ``"Master price table_DT_CHICONY"``  →  ``"CHICONY"``
    """
    # Split on first underscore: "Master price table" + "NB_Chicony"
    parts = folder_name.split("_", 1)
    if len(parts) < 2:
        return folder_name
    rest = parts[1]          # "NB_Chicony"
    # Split on first underscore: "NB" + "Chicony"
    parts2 = rest.split("_", 1)
    return parts2[1] if len(parts2) > 1 else rest


def _find_segment_root(parent: Path, keyword: str) -> Path | None:
    """Find a sub-folder whose name contains *keyword* (case-insensitive)."""
    keyword_lower = keyword.lower()
    for sub in parent.iterdir():
        if sub.is_dir() and keyword_lower in sub.name.lower():
            return sub
    return None


def _fix_gtk_suppliers(wb: Workbook, supplier_name: str) -> None:
    """Overwrite every 'GTK Suppliers' column (Row 3+) with *supplier_name*."""
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col_idx)
            if cell.value is None:
                continue
            normalised = str(cell.value).replace("\n", " ").strip().lower()
            if normalised == "gtk suppliers":
                for row_idx in range(3, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).value = supplier_name


def _copy_ws_to_wb(src_ws, target_wb: Workbook, sheet_name: str) -> None:
    """Copy all rows from *src_ws* into a new sheet in *target_wb*."""
    new_ws = target_wb.create_sheet(sheet_name)
    for row in src_ws.iter_rows(values_only=True):
        new_ws.append(list(row))


# ---------------------------------------------------------------------------
# Public functions
# ---------------------------------------------------------------------------

def ingest_nb(
    nb_kb_path: Path,
    raw_nb_dir: Path,
    log: Callable[[str, str], None],
) -> dict[str, list[str]]:
    """Stage 1 for NB KB — split each supplier into bNB and cNB workbooks.

    *raw_nb_dir* is ``source data/NB``; outputs go to
    ``raw_nb_dir/bNB`` and ``raw_nb_dir/cNB``.

    Returns a dict ``{"bNB": [supplier, ...], "cNB": [supplier, ...]}``.
    """
    seg_root = _find_segment_root(nb_kb_path, "Master price table_NB")
    if seg_root is None:
        log(f"Cannot find 'Master price table_NB' sub-folder in {nb_kb_path}", "ERROR")
        return {"bNB": [], "cNB": []}

    # -----------------------------------------------------------------------
    # Collect all supplier workbooks (apply GTK fix in-memory)
    # -----------------------------------------------------------------------
    nb_folder_re = _make_supplier_folder_re("NB")
    supplier_data: dict[str, dict] = {}
    for supplier_folder in sorted(seg_root.iterdir()):
        if not supplier_folder.is_dir():
            continue
        if not nb_folder_re.match(supplier_folder.name):
            log(
                f"  NB: skipping '{supplier_folder.name}' "
                f"(must match 'Master price table_NB_<Supplier>')",
                "WARNING",
            )
            continue
        latest = _get_latest_xlsx(supplier_folder)
        if latest is None:
            log(f"  NB: no .xlsx found in '{supplier_folder.name}', skipping", "WARNING")
            continue

        supplier_name = _parse_supplier_name(supplier_folder.name)
        log(f"  NB supplier: {supplier_name}  ←  {latest.name}", "INFO")

        wb = load_workbook(latest, data_only=True)
        _fix_gtk_suppliers(wb, supplier_name)
        supplier_data[supplier_name] = {"path": latest, "wb": wb}

    if not supplier_data:
        log("  NB: no supplier data found", "WARNING")
        return {"bNB": [], "cNB": []}

    # -----------------------------------------------------------------------
    # Pass 1 — find FY years that have BOTH cNB and bNB per supplier
    # (each supplier is evaluated independently; no cross-supplier intersection)
    # -----------------------------------------------------------------------
    valid_fy_per_supplier: dict[str, set[str]] = {}
    for supplier_name, data in supplier_data.items():
        fy_types: dict[str, set[str]] = {}
        for ws in data["wb"].worksheets:
            m = NB_SHEET_RE.match(ws.title.strip())
            if m:
                fy = f"FY{m.group(1)}"
                nb_type = m.group(2).lower()
                fy_types.setdefault(fy, set()).add(nb_type)
        # Only include FY years where THIS supplier has both 'b' and 'c'
        both = {fy for fy, types in fy_types.items() if "c" in types and "b" in types}
        valid_fy_per_supplier[supplier_name] = both
        log(f"  NB {supplier_name} valid FY: {sorted(both)}", "INFO")

    # -----------------------------------------------------------------------
    # Pass 2 — combine bNB + cNB into one workbook per supplier
    # -----------------------------------------------------------------------
    raw_nb_dir.mkdir(parents=True, exist_ok=True)

    coverage_bnb: dict[str, list[str]] = {}
    coverage_cnb: dict[str, list[str]] = {}
    for supplier_name, data in supplier_data.items():
        supplier_valid_fy = valid_fy_per_supplier.get(supplier_name, set())
        wb_out: Workbook = Workbook()
        wb_out.remove(wb_out.active)  # type: ignore[arg-type]

        fy_bnb: list[str] = []
        fy_cnb: list[str] = []

        for ws in data["wb"].worksheets:
            m = NB_SHEET_RE.match(ws.title.strip())
            if not m:
                continue
            fy = f"FY{m.group(1)}"
            if fy not in supplier_valid_fy:
                continue
            nb_type = m.group(2).lower()
            sheet_name = f"{fy} bNB" if nb_type == "b" else f"{fy} cNB"
            _copy_ws_to_wb(ws, wb_out, sheet_name)
            if nb_type == "b":
                fy_bnb.append(fy)
            else:
                fy_cnb.append(fy)

        if wb_out.worksheets:
            fname = f"{data['path'].stem}.xlsx"
            wb_out.save(raw_nb_dir / fname)
            log(
                f"    → NB/{fname}  "
                f"({len(fy_bnb)} bNB + {len(fy_cnb)} cNB sheets)",
                "INFO",
            )
            if fy_bnb:
                coverage_bnb[supplier_name] = fy_bnb
            if fy_cnb:
                coverage_cnb[supplier_name] = fy_cnb

    return {"bNB": coverage_bnb, "cNB": coverage_cnb}


def ingest_segment(
    segment_path: Path,
    segment_name: str,
    raw_dir: Path,
    log: Callable[[str, str], None],
) -> list[str]:
    """Stage 1 for DT or Peripheral.

    *raw_dir* is the destination folder (e.g. ``source data/DT``).
    *segment_name* should be ``"DT"`` or ``"Peripheral"``.
    Returns ``{supplier_name: [FY sheets]}`` for processed suppliers.
    """
    seg_root = _find_segment_root(
        segment_path, f"Master price table_{segment_name}"
    )
    if seg_root is None:
        log(
            f"Cannot find 'Master price table_{segment_name}' sub-folder "
            f"in {segment_path}",
            "ERROR",
        )
        return {}

    raw_dir.mkdir(parents=True, exist_ok=True)

    supplier_folder_re = _make_supplier_folder_re(segment_name)
    coverage: dict[str, list[str]] = {}
    for supplier_folder in sorted(seg_root.iterdir()):
        if not supplier_folder.is_dir():
            continue
        if not supplier_folder_re.match(supplier_folder.name):
            log(
                f"  {segment_name}: skipping '{supplier_folder.name}' "
                f"(must match 'Master price table_{segment_name}_<Supplier>')",
                "WARNING",
            )
            continue
        latest = _get_latest_xlsx(supplier_folder)
        if latest is None:
            log(f"  {segment_name}: no .xlsx in '{supplier_folder.name}', skipping", "WARNING")
            continue

        supplier_name = _parse_supplier_name(supplier_folder.name)
        log(f"  {segment_name} supplier: {supplier_name}  ←  {latest.name}", "INFO")

        wb = load_workbook(latest, data_only=True)
        _fix_gtk_suppliers(wb, supplier_name)

        # Build output workbook with only FY sheets, normalised names
        wb_out: Workbook = Workbook()
        wb_out.remove(wb_out.active)  # type: ignore[arg-type]

        for ws in wb.worksheets:
            stripped = ws.title.strip()
            if FY_SHEET_RE.match(stripped):
                _copy_ws_to_wb(ws, wb_out, stripped.upper())

        if wb_out.worksheets:
            wb_out.save(raw_dir / latest.name)
            log(f"    → {segment_name}/{latest.name}  ({len(wb_out.worksheets)} sheets)", "INFO")
            coverage[supplier_name] = [ws.title for ws in wb_out.worksheets]

    return coverage
